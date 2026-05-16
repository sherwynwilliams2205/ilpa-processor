"""
Layout Learner
Given a reference PDF and a solved ILPA Excel template,
reverse-engineers the field → row mapping automatically.

Returns a structured dict that codegen.py turns into a layout_X.py file.
"""

import openpyxl
from .extractor import extract

TARGET_ROWS = range(9, 91)          # rows that carry data in the template
COL_MAP     = {"E": 5, "F": 6, "G": 7}
COL_HINT    = {5: 0, 6: 1, 7: 2}   # Excel column index → PDF list position hint


def learn(pdf_path: str, excel_path: str) -> dict:
    """
    Returns:
      {
        "row_map": {
            row_int: {
                col_letter: {
                    "type":   "pdf" | "fixed" | "zero",
                    # pdf entries also have:
                    "field":  str,
                    "col":    int,   # 0=QTD, 1=YTD, 2=SI
                    "negate": bool,
                    "value":  float,
                }
            }
        },
        "pdf_numbers": { field: [QTD, YTD, SI] },
        "stats": {"matched": int, "fixed": int, "zero": int},
      }
    """
    raw         = extract(pdf_path)
    pdf_numbers = raw["numbers"]

    # Build value lookup: rounded_abs_value → [(field, col_idx, raw_value)]
    val_lookup: dict[int, list] = {}
    for field, vals in pdf_numbers.items():
        for i, v in enumerate(vals[:3]):
            if v is not None and abs(v) > 0:
                av = round(abs(v))
                val_lookup.setdefault(av, []).append((field, i, v))

    # Load solved Excel (data_only so we get values, not formula strings)
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = ("Reporting Template"
                  if "Reporting Template" in wb.sheetnames
                  else wb.sheetnames[0])
    ws = wb[sheet_name]

    row_map = {}
    stats   = {"matched": 0, "fixed": 0, "zero": 0}

    for row in TARGET_ROWS:
        cols = {}
        for col_letter, col_idx in COL_MAP.items():
            val = ws.cell(row=row, column=col_idx).value

            # Skip blanks and formula strings
            if val is None or isinstance(val, str):
                continue

            if val == 0:
                cols[col_letter] = {"type": "zero"}
                stats["zero"] += 1
                continue

            av    = round(abs(val))
            hint  = COL_HINT[col_idx]
            match = _find_match(av, hint, val_lookup)

            if match:
                field, pdf_col, pdf_val = match
                negate = (val < 0) != (pdf_val < 0)
                cols[col_letter] = {
                    "type":   "pdf",
                    "field":  field,
                    "col":    pdf_col,
                    "negate": negate,
                    "value":  val,
                }
                stats["matched"] += 1
            else:
                cols[col_letter] = {"type": "fixed", "value": val}
                stats["fixed"]   += 1

        if cols:
            row_map[row] = cols

    return {
        "row_map":     row_map,
        "pdf_numbers": pdf_numbers,
        "stats":       stats,
    }


def _find_match(av: int, col_hint: int,
                val_lookup: dict) -> tuple | None:
    """
    Find the best matching PDF field for a given absolute cell value.
    Prefers column-position match.  Tolerance: 0.5% or $1, whichever larger.
    """
    tolerance  = max(1, av * 0.005)
    best       = None
    best_score = -1.0

    for k, entries in val_lookup.items():
        diff = abs(k - av)
        if diff > tolerance:
            continue
        for entry in entries:
            _, pdf_col, _ = entry
            score  = (10.0 if pdf_col == col_hint else 5.0)
            score -= diff / (av + 1)
            if score > best_score:
                best_score = score
                best       = entry

    return best
