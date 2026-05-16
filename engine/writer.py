"""
Template Writer
Copies the master template and populates INPUT cells from a row_data dict.
Leaves all formula cells untouched.
Writes period dates to P2:P5 from the quarter string.
"""

import shutil
import openpyxl
from datetime import date, timedelta
from pathlib import Path

TEMPLATE_PATH = Path(__file__).parent.parent / "assets" / "ILPA_Master_Template.xlsx"

COL = {"E": 5, "F": 6, "G": 7,
       "H": 8, "I": 9, "J": 10,
       "K": 11, "L": 12, "M": 13}

# Known inception dates per layout / fund
INCEPTION_DATES = {
    "B": date(2007, 6, 21),   # Kelso Investment Associates VIII — June 21 2007
    "A": None,                 # Veritas — populate when known
}

# Quarter → (period_start_month, period_end_month, period_end_day)
QUARTER_MAP = {
    "Q1": (1,  3,  31),
    "Q2": (4,  6,  30),
    "Q3": (7,  9,  30),
    "Q4": (10, 12, 31),
}


def _quarter_dates(quarter: str, layout: str):
    """
    Parse "Q2_2022" → (inception, year_start, period_start, period_end).
    Returns date objects (or None for inception if unknown).
    """
    try:
        q, yr = quarter.split("_")
        year = int(yr)
    except Exception:
        return None, None, None, None

    start_m, end_m, end_d = QUARTER_MAP.get(q, (1, 3, 31))

    inception    = INCEPTION_DATES.get(layout)
    year_start   = date(year, 1, 1)
    period_start = date(year, start_m, 1)
    period_end   = date(year, end_m, end_d)

    return inception, year_start, period_start, period_end


def write(row_data: dict, output_path: str,
          fund_name: str = "", layout: str = "",
          quarter: str = "") -> str:
    """
    row_data  : {row_int: {"E": val, "F": val, "G": val}}
    None values = leave cell alone (formula stays).
    Returns output_path.
    """
    shutil.copy2(TEMPLATE_PATH, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Reporting Template"]

    # ── Fund name → B3 ──────────────────────────────────────────
    if fund_name:
        try:
            ws["B3"] = fund_name
        except AttributeError:
            pass

    # ── Period dates → P2:P5 ────────────────────────────────────
    if quarter:
        inception, year_start, period_start, period_end = _quarter_dates(quarter, layout)
        date_map = {
            "P2": inception,      # Inception Start
            "P3": year_start,     # Current Year Start (Jan 1)
            "P4": period_start,   # Current Period Start
            "P5": period_end,     # Period End
        }
        for cell_ref, dt in date_map.items():
            if dt is not None:
                try:
                    ws[cell_ref] = dt
                    ws[cell_ref].number_format = "dd/mm/yyyy"
                except AttributeError:
                    pass

    # ── Row data → E:M columns ──────────────────────────────────
    for row_num, cols in row_data.items():
        for col_letter, value in cols.items():
            if value is None:
                continue
            col_idx = COL.get(col_letter)
            if col_idx is None:
                continue
            try:
                ws.cell(row=row_num, column=col_idx).value = value
            except AttributeError:
                pass   # merged cell slave — skip

    wb.save(output_path)
    return output_path
